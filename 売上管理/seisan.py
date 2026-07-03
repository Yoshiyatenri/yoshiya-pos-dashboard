"""
精算処理スクリプト (seisan.py)
Power Automate Desktop フローを Python で再現

必要パッケージ:
  pip install openpyxl selenium pywin32
"""
import csv
import json
import sys
import time
from datetime import date, timedelta
from pathlib import Path
import tkinter as tk
from tkinter import messagebox, simpledialog

# ─── パッケージ確認 ───────────────────────────────────────────
try:
    import openpyxl
except ImportError:
    print("openpyxl が必要です: pip install openpyxl"); sys.exit(1)

try:
    from selenium import webdriver
    from selenium.webdriver.edge.options import Options as EdgeOptions
    from selenium.webdriver.common.by import By
    from selenium.webdriver.common.keys import Keys
    from selenium.webdriver.common.action_chains import ActionChains
except ImportError:
    print("selenium が必要です: pip install selenium"); sys.exit(1)

try:
    import win32com.client
except ImportError:
    print("pywin32 が必要です: pip install pywin32"); sys.exit(1)

# ─── 設定 ─────────────────────────────────────────────────────
BASE_DIR = Path(__file__).parent
CONFIG_PATH = BASE_DIR / "seisan_config.json"

def load_config():
    with open(CONFIG_PATH, encoding="utf-8") as f:
        return json.load(f)

# ─── 定数：レジ構成 ────────────────────────────────────────────
# 店舗コード → レジ表示名リスト
REGISTER_NAMES = {
    1: ["1レジ", "2レジ"],
    6: ["1レジ", "2レジ", "3レジ", "4レジ", "5レジ", "6レジ"],
    7: ["1レジ", "2レジ", "3レジ", "11レジ", "12レジ"],
    8: ["1レジ", "2レジ", "3レジ", "11レジ", "12レジ"],
    9: ["1レジ", "2レジ", "3レジ", "4レジ"],
}

# 店舗コード → ExcelInstance4 ループ上限 (Power Automate の NewVar2)
LOOP_END = {7: 7, 8: 7, 9: 6, 6: 8}

def get_reg_defaults(store_code, count1):
    """1レジ精算回数を元に各レジのデフォルト値リスト (None=入力必須)"""
    n = count1
    if store_code == 9:   return [None, None, None, None]
    elif store_code == 6: return [None, n - 1, n - 1, n - 2, None, None]
    elif store_code in (7, 8): return [None, None, None, None, None]
    else:                 return [None, None]


# ─── キャンセル例外 ───────────────────────────────────────────
class CancelError(Exception):
    pass


# ─── GUI ダイアログ ────────────────────────────────────────────
class Dialogs:
    """tkinter ベースの入力ダイアログ群"""

    def __init__(self):
        self.root = tk.Tk()
        self.root.withdraw()

    def destroy(self):
        try:
            self.root.destroy()
        except Exception:
            pass

    def ask_store(self, stores):
        """店舗選択 → 店舗名を返す"""
        dlg = tk.Toplevel(self.root)
        dlg.title("店舗選択"); dlg.resizable(False, False); dlg.grab_set()
        tk.Label(dlg, text="店舗を選択してください", padx=10, pady=8).pack()
        lb = tk.Listbox(dlg, selectmode=tk.SINGLE, width=25, height=len(stores))
        for s in stores: lb.insert(tk.END, s)
        lb.select_set(0); lb.pack(padx=15, pady=4)
        result = [None]

        def ok():
            sel = lb.curselection()
            if sel: result[0] = stores[sel[0]]; dlg.destroy()
        def cancel(): dlg.destroy()

        bf = tk.Frame(dlg); bf.pack(pady=6)
        tk.Button(bf, text="OK", command=ok, width=10).pack(side=tk.LEFT, padx=4)
        tk.Button(bf, text="キャンセル", command=cancel, width=10).pack(side=tk.LEFT, padx=4)
        dlg.bind("<Return>", lambda e: ok()); dlg.wait_window()
        if result[0] is None: raise CancelError
        return result[0]

    def ask_date(self, default: date):
        """日付入力 → date を返す"""
        dlg = tk.Toplevel(self.root)
        dlg.title("売上日入力"); dlg.resizable(False, False); dlg.grab_set()
        tk.Label(dlg, text="売上日を入力してください", padx=10, pady=8).pack()
        frame = tk.Frame(dlg); frame.pack(padx=15, pady=4)
        year_v  = tk.StringVar(value=str(default.year))
        month_v = tk.StringVar(value=f"{default.month:02d}")
        day_v   = tk.StringVar(value=f"{default.day:02d}")
        tk.Label(frame, text="年").grid(row=0, column=0, padx=2)
        tk.Entry(frame, textvariable=year_v, width=6).grid(row=0, column=1)
        tk.Label(frame, text="月").grid(row=0, column=2, padx=2)
        tk.Entry(frame, textvariable=month_v, width=4).grid(row=0, column=3)
        tk.Label(frame, text="日").grid(row=0, column=4, padx=2)
        tk.Entry(frame, textvariable=day_v, width=4).grid(row=0, column=5)
        result = [None]

        def ok():
            try:
                d = date(int(year_v.get()), int(month_v.get()), int(day_v.get()))
                result[0] = d; dlg.destroy()
            except ValueError:
                messagebox.showerror("エラー", "日付の形式が正しくありません", parent=dlg)
        def cancel(): dlg.destroy()

        bf = tk.Frame(dlg); bf.pack(pady=6)
        tk.Button(bf, text="OK", command=ok, width=10).pack(side=tk.LEFT, padx=4)
        tk.Button(bf, text="キャンセル", command=cancel, width=10).pack(side=tk.LEFT, padx=4)
        dlg.bind("<Return>", lambda e: ok()); dlg.wait_window()
        if result[0] is None: raise CancelError
        return result[0]

    def ask_text(self, title, message):
        val = simpledialog.askstring(title, message, parent=self.root)
        if val is None: raise CancelError
        return val

    def ask_number(self, title, message, default=None):
        default_str = str(int(default)) if default is not None else ""
        while True:
            val = simpledialog.askstring(title, message, initialvalue=default_str, parent=self.root)
            if val is None: raise CancelError
            try: return int(val)
            except ValueError:
                messagebox.showerror("エラー", "数値を入力してください", parent=self.root)

    def ask_registers(self, store_code):
        """レジ精算回数を一画面でまとめて入力 → [int, ...] を返す"""
        reg_names = REGISTER_NAMES[store_code]
        dlg = tk.Toplevel(self.root)
        dlg.title("レジ精算回数入力"); dlg.resizable(False, False); dlg.grab_set()
        tk.Label(dlg, text="各レジの精算回数を入力してください", padx=10, pady=8).pack()
        frame = tk.Frame(dlg); frame.pack(padx=20, pady=4)
        entries = []
        for i, name in enumerate(reg_names):
            tk.Label(frame, text=f"{name}精算回数:", anchor="e", width=14).grid(row=i, column=0, pady=2, padx=4)
            e = tk.Entry(frame, width=8); e.grid(row=i, column=1, pady=2); entries.append(e)

        def on_first_leave(event):
            try:
                n = int(entries[0].get())
                defaults = get_reg_defaults(store_code, n)
                for i, e in enumerate(entries[1:], 1):
                    if not e.get() and defaults[i] is not None:
                        e.delete(0, tk.END); e.insert(0, str(defaults[i]))
            except ValueError: pass
        entries[0].bind("<FocusOut>", on_first_leave)

        result = [None]
        def ok():
            try:
                vals = [int(e.get()) for e in entries]
                result[0] = vals; dlg.destroy()
            except ValueError:
                messagebox.showerror("エラー", "すべてに数値を入力してください", parent=dlg)
        def cancel(): dlg.destroy()

        bf = tk.Frame(dlg); bf.pack(pady=6)
        tk.Button(bf, text="OK", command=ok, width=10).pack(side=tk.LEFT, padx=4)
        tk.Button(bf, text="キャンセル", command=cancel, width=10).pack(side=tk.LEFT, padx=4)
        dlg.bind("<Return>", lambda e: ok()); entries[0].focus_set(); dlg.wait_window()
        if result[0] is None: raise CancelError
        return result[0]

    def show_cash_confirm(self, cash_amount):
        """現金確認ダイアログ → True=OK"""
        dlg = tk.Toplevel(self.root)
        dlg.title("現金を数えて確認してください。"); dlg.resizable(False, False); dlg.grab_set()
        tk.Label(dlg, text=f"現金　{cash_amount}　円\n\n合っていればOKを押してください。",
                 padx=20, pady=10, justify=tk.LEFT).pack()
        result = [False]
        def ok(): result[0] = True; dlg.destroy()
        def cancel(): dlg.destroy()
        bf = tk.Frame(dlg); bf.pack(pady=6)
        tk.Button(bf, text="OK", command=ok, width=10).pack(side=tk.LEFT, padx=4)
        tk.Button(bf, text="キャンセル", command=cancel, width=10).pack(side=tk.LEFT, padx=4)
        dlg.wait_window(); return result[0]


# ─── Selenium: CSV ダウンロード ────────────────────────────────
def _build_driver(download_dir: str) -> webdriver.Edge:
    opts = EdgeOptions()
    opts.add_experimental_option("prefs", {
        "download.default_directory": download_dir,
        "download.prompt_for_download": False,
        "download.directory_upgrade": True,
        "safebrowsing.enabled": True,
    })
    opts.add_argument("--disable-notifications")
    opts.add_argument("--start-maximized")
    return webdriver.Edge(options=opts)


def _login(driver, cfg):
    driver.get(cfg["netdoa_login_url"])
    time.sleep(3)
    try:
        driver.find_element(By.NAME, "ws_pswd").send_keys(cfg["password"])
        time.sleep(1)
        driver.find_element(By.NAME, "ws_userId").send_keys(cfg["user_id"])
        time.sleep(1)
        driver.execute_script("document.Login.submit()")
    except Exception:
        try:
            driver.find_element(By.NAME, "ws_userId").send_keys(cfg["user_id"])
            driver.find_element(By.NAME, "ws_pswd").send_keys(cfg["password"])
            driver.find_element(By.NAME, "ws_pswd").send_keys(Keys.RETURN)
        except Exception as e:
            print(f"ログイン操作エラー: {e}")
    time.sleep(5)


def _click_csv_button(driver, tab_count: int):
    """CSVボタンをクリック（JS検索 → Tabフォールバック）"""
    found = driver.execute_script("""
        var els = document.querySelectorAll('input[type=button],input[type=submit],button,a');
        for(var el of els){
            var t = (el.value || el.textContent || '');
            if(t.indexOf('CSV') >= 0 || t.indexOf('ＣＳＶ') >= 0){
                el.click(); return true;
            }
        }
        return false;
    """)
    if not found:
        body = driver.find_element(By.TAG_NAME, "body")
        body.click()
        for _ in range(tab_count):
            ActionChains(driver).send_keys(Keys.TAB).perform()
            time.sleep(0.08)
        ActionChains(driver).send_keys(Keys.RETURN).perform()


def _click_tab_text(driver, text):
    """テキストでタブ/リンクを探してクリック"""
    driver.execute_script(f"""
        var els = document.querySelectorAll('a,input,button,td,th,span,div');
        for(var el of els){{
            if((el.textContent||el.value||'').trim().indexOf('{text}') >= 0){{
                el.click(); return;
            }}
        }}
    """)


def download_csvs(cfg, store_name, store_code, sale_date):
    """
    Netdoa NX にログインして 4 つの CSV をダウンロードする。
    """
    yr  = sale_date.strftime("%Y")
    mm  = sale_date.strftime("%m")
    dd  = sale_date.strftime("%d")
    sc  = store_code
    dl  = Path(cfg["download_dir"])

    base = "https://www.netdoa-nx.jp"
    qp   = (f"ws_bussId=7&ws_streFlg=3&ws_streCd={sc}&"
            f"ws_yearFrom={yr}&ws_monthFrom={mm}&ws_dayFrom={dd}&"
            f"ws_yearTo={yr}&ws_monthTo={mm}&ws_dayTo={dd}")

    sar6_1 = (f"{base}/SA/SAR6/SAR60010010.php?{qp}&"
              f"ws_programID=SAR6&ws_actType=1&ws_clsFlg=3&ws_clsCd=0&"
              f"ws_makerCd=&ws_catDscCd=0&ws_smlClsCd=0&ws_periodFlg=2&"
              f"ws_planCd=&ws_disp_yearFrom={yr}&ws_disp_monthFrom={mm}&"
              f"ws_disp_dayFrom=&ws_disp_yearTo={yr}&ws_disp_monthTo={mm}&"
              f"ws_disp_dayTo=&ws_strePermit=0&ws_clsPermit=0&ws_periodPermit=&ws_SerialNo=1")
    sar6_0 = sar6_1.replace("ws_SerialNo=1", "ws_SerialNo=0")

    sats = (f"{base}/SA/SATS/SATS0010010.php?{qp}&"
            f"ws_programID=SATS&ws_actType=1&ws_clsFlg=2&ws_clsCd=999999&"
            f"ws_makerCd=&ws_catDscCd=0&ws_smlClsCd=0&ws_periodFlg=2&"
            f"ws_planCd=&ws_strePermit=0&ws_clsPermit=&ws_periodPermit=&ws_SerialNo=")

    sacl = (f"{base}/SA/SACL/SACL0010010.php?{qp}&"
            f"ws_programID=SACL&ws_actType=1&ws_clsFlg=1&ws_clsCd=0&"
            f"ws_makerCd=&ws_catDscCd=0&ws_smlClsCd=0&ws_periodFlg=2&"
            f"ws_planCd=&ws_disp_yearFrom={yr}&ws_disp_monthFrom={mm}&"
            f"ws_disp_dayFrom=&ws_disp_yearTo={yr}&ws_disp_monthTo={mm}&"
            f"ws_disp_dayTo=&ws_strePermit=0&ws_clsPermit=0&ws_periodPermit=&ws_SerialNo=")

    driver = _build_driver(str(dl))
    try:
        _login(driver, cfg)

        print("1/4 中分類別分類売上一覧 ダウンロード中...")
        driver.get(sar6_1); time.sleep(4)
        _click_csv_button(driver, 19)
        time.sleep(10)
        ActionChains(driver).send_keys(Keys.TAB).send_keys(Keys.RETURN).perform()
        time.sleep(3)

        print("2/4 SAR6 SerialNo=0 ダウンロード中...")
        driver.get(sar6_0); time.sleep(6)
        _click_csv_button(driver, 19)
        time.sleep(8)
        ActionChains(driver).send_keys(Keys.TAB).send_keys(Keys.RETURN).perform()
        time.sleep(3)

        print("3/4 大分類別時間帯販売 ダウンロード中...")
        driver.get(sats); time.sleep(6)
        _click_csv_button(driver, 25)
        time.sleep(6)
        ActionChains(driver).send_keys(Keys.TAB).send_keys(Keys.RETURN).perform()
        time.sleep(4)

        print("4/4 取引レポート (レジ別・店舗別) ダウンロード中...")
        driver.get(sacl); time.sleep(5)
        # レジ別タブをクリック
        _click_tab_text(driver, "レジ別"); time.sleep(5)
        _click_csv_button(driver, 15)
        time.sleep(6)
        ActionChains(driver).send_keys(Keys.TAB).send_keys(Keys.RETURN).perform()
        time.sleep(6)
        # 店舗別タブをクリック
        _click_tab_text(driver, "店舗別"); time.sleep(3)
        body = driver.find_element(By.TAG_NAME, "body"); body.click()
        ActionChains(driver).key_down(Keys.SHIFT)\
            .send_keys(Keys.TAB).send_keys(Keys.TAB)\
            .key_up(Keys.SHIFT).send_keys(Keys.RETURN).perform()
        time.sleep(6)
        # 追加の保存確認操作
        for _ in range(3):
            ActionChains(driver).send_keys(Keys.TAB).perform(); time.sleep(2)
            ActionChains(driver).send_keys(Keys.RETURN).perform(); time.sleep(2)

    finally:
        driver.quit()

    print("ダウンロード完了")


# ─── CSV ヘルパー ──────────────────────────────────────────────
def read_csv_sj(path):
    """Shift-JIS CSV を 2D リスト (0-indexed) で読む"""
    with open(path, encoding="shift_jis", errors="replace", newline="") as f:
        return list(csv.reader(f))


def csv_val(rows, row_1, col_1):
    """1-indexed 行列でセル値取得。数値文字列は数値変換する"""
    try:
        v = rows[row_1 - 1][col_1 - 1].replace(",", "").strip()
    except IndexError:
        return ""
    if not v: return ""
    try:   return int(v)
    except ValueError:
        try: return float(v)
        except ValueError: return v


# ─── Excel ヘルパー ────────────────────────────────────────────
def find_col(ws, search_row, header_text):
    """
    指定行でヘッダーテキストを検索し Excel 列番号(1-indexed)を返す。
    Power Automate の DataTable 列インデックス = 戻り値 - 1
    """
    for col in range(1, 300):
        v = ws.cell(row=search_row, column=col).value
        if v is not None and str(v).strip() == header_text:
            return col  # Excel 列番号 (1-indexed)
    return None


def find_row(ws, search_val):
    """
    シート全体を列→行の順で検索し、値と一致する行番号(1-indexed)を返す。
    Power Automate の SearchBy: Columns に対応。
    """
    max_col = ws.max_column or 100
    max_row = ws.max_row or 100
    for col in range(1, max_col + 1):
        for row in range(1, max_row + 1):
            v = ws.cell(row=row, column=col).value
            try:
                if int(v) == int(search_val):
                    return row
            except (TypeError, ValueError):
                pass
    return None


def xw(ws, row, excel_col, value):
    """Excel 列番号(1-indexed)でセルに書き込む"""
    ws.cell(row=row, column=excel_col).value = value


# ─── メイン処理 ────────────────────────────────────────────────
def process_seisan(dlg, cfg, store_name, store_code, sale_date,
                   operator, reg_counts, weather, temp_max, temp_min):
    yr   = sale_date.year
    mm   = sale_date.month
    dd   = sale_date.day
    mm2  = f"{mm:02d}"
    dd2  = f"{dd:02d}"
    yy   = yr % 100     # 年下2桁

    dl_dir = Path(cfg["download_dir"])
    filemei1 = dl_dir / f"{store_name}中分類別分類売上一覧 (日：{yr}.{mm2}.{dd2}～{yr}.{mm2}.{dd2}).csv"
    filemei2 = dl_dir / f"店舗：{store_code} 取引レポート (レジ別) 【ＣＳＶ】  [{yr}.{mm2}.{dd2} ～ {yr}.{mm2}.{dd2}].csv"
    filemei3 = dl_dir / f"{store_name}大分類別時間帯販売 (日：{yr}.{mm2}.{dd2}～{yr}.{mm2}.{dd2}).csv"
    filemei4 = dl_dir / f"店舗：{store_code} 取引レポート (店舗別) 【ＣＳＶ】  [{yr}.{mm2}.{dd2} ～ {yr}.{mm2}.{dd2}].csv"

    # ── 売上表 Excel ──────────────────────────────────────────
    sales_path = Path(cfg["sales_table"].replace("{store}", store_name))
    print(f"売上表 を開いています: {sales_path.name}")
    wb_sales = openpyxl.load_workbook(str(sales_path))
    sheet_idx = mm - 6 if mm >= 7 else mm + 6  # 7月=1, 8月=2, ... 6月=12
    ws = wb_sales.worksheets[sheet_idx - 1]

    # 4行目からヘッダー列位置を取得 (Excel 列番号, 1-indexed)
    # Power Automate の paypayCOL[0][1] = find_col("PayPay") - 1
    kion_col      = find_col(ws, 4, "気温")
    tensu_col     = find_col(ws, 4, "点数")
    tenki_col     = find_col(ws, 4, "天気")
    syuuryou_col  = find_col(ws, 4, "終了")
    kyaku_col     = find_col(ws, 4, "客数")
    uri_col       = find_col(ws, 4, "売上")
    paypay_col    = find_col(ws, 4, "PayPay")
    cashless_col  = find_col(ws, 4, "キャッシュレス合計")
    tokubai_col   = find_col(ws, 4, "カテゴリー")
    kaiten_col    = find_col(ws, 4, "開店後1時間売上")
    heiten_col    = find_col(ws, 4, "閉店前1時間売上")

    # 日付行を検索
    excel_row = find_row(ws, dd)
    print(f"売上日 {dd}日 → 行 {excel_row}")

    # ── filemei1: カテゴリー別売上を転記 ──────────────────────
    csv1 = read_csv_sj(filemei1)
    # tokubaiCOL[0][1]+LoopIndex+1 = (tokubai_col-1)+LoopIndex+1 = tokubai_col+LoopIndex
    # LoopIndex FROM 1 TO (syuuryou_col-1)-(tokubai_col-1)+1 = syuuryou_col-tokubai_col+1
    loop_max = syuuryou_col - tokubai_col + 1
    for loop_idx in range(1, loop_max + 1):
        read_col  = tokubai_col + loop_idx  # Excel 列 (3行目のカテゴリー名)
        write_col = tokubai_col + loop_idx  # 書き込み先 Excel 列 (同じ)
        category  = ws.cell(row=3, column=read_col).value
        if not category or str(category).strip() == "":
            continue
        # filemei1 でカテゴリー名を検索して col3 の値を取得
        found_row = None
        for r_idx, row_data in enumerate(csv1):
            if any(str(c).strip() == str(category).strip() for c in row_data):
                found_row = r_idx + 1  # 1-indexed
                break
        if found_row is None:
            continue
        try:
            val = csv_val(csv1, found_row, 3)
            xw(ws, excel_row, write_col, val)
        except Exception:
            pass  # ON ERROR 相当

    # ── filemei3: 開店・閉店1時間データ ──────────────────────
    csv3 = read_csv_sj(filemei3)
    last_r = len(csv3)
    heiten_uri    = csv_val(csv3, last_r, 2)
    heiten_ninzu  = csv_val(csv3, last_r, 4)
    kaiten_uri    = csv_val(csv3, 2, 2)
    kaiten_ninzu  = csv_val(csv3, 2, 4)

    # heitenCOL[0][1]+1 = heiten_col  /  +2 = heiten_col+1
    xw(ws, excel_row, heiten_col,     heiten_uri)
    xw(ws, excel_row, heiten_col + 1, heiten_ninzu)
    xw(ws, excel_row, kaiten_col,     kaiten_uri)
    xw(ws, excel_row, kaiten_col + 1, kaiten_ninzu)

    # ── filemei2: 取引レポート (レジ別) ─────────────────────
    csv2 = read_csv_sj(filemei2)
    loop_end = LOOP_END.get(store_code, 4)

    daily_path = (Path(cfg["daily_report_dir"])
                  / f"{yy}.{mm2}"
                  / f"082天理）売上日報 {yy}.{mm2}.xlsx")
    print(f"日報 を開いています: {daily_path.name}")
    wb_daily = openpyxl.load_workbook(str(daily_path))
    ws_daily = wb_daily.worksheets[dd - 1]  # ActivateWorksheetByIndex(daynum) = 0-indexed dd-1

    # 日報ヘッダー (行60)
    ws_daily.cell(row=60, column=2).value  = yr
    ws_daily.cell(row=60, column=4).value  = f"{mm}/{dd}"
    ws_daily.cell(row=60, column=8).value  = operator
    ws_daily.cell(row=60, column=10).value = "とれとれ市場" if store_code == 7 else store_name

    # レジ別ループ (LoopIndex FROM 3 TO NewVar2)
    for loop_idx in range(3, loop_end + 1):
        kyakusu    = csv_val(csv2, 6,   loop_idx)
        tensuu     = csv_val(csv2, 8,   loop_idx)
        genuri1    = csv_val(csv2, 18,  loop_idx)
        kinken     = csv_val(csv2, 118, loop_idx)
        paypayuri  = csv_val(csv2, 25,  loop_idx)
        imanarauri = csv_val(csv2, 28,  loop_idx)
        ichikauri  = csv_val(csv2, 31,  loop_idx)
        credituri  = csv_val(csv2, 34,  loop_idx)
        iduri      = csv_val(csv2, 37,  loop_idx)
        koutuuri   = csv_val(csv2, 40,  loop_idx)
        edyuri     = csv_val(csv2, 43,  loop_idx)
        waonuri    = csv_val(csv2, 46,  loop_idx)
        nanacouri  = csv_val(csv2, 49,  loop_idx)
        ginrenuri  = csv_val(csv2, 52,  loop_idx)
        uri11      = csv_val(csv2, 55,  loop_idx)
        uri13      = csv_val(csv2, 79,  loop_idx)

        row_b = loop_idx + 62  # 日報の書き込み行
        seisansuu = reg_counts[loop_idx - 3] if (loop_idx - 3) < len(reg_counts) else ""

        ws_daily.cell(row=row_b, column=5).value  = kyakusu
        ws_daily.cell(row=row_b, column=3).value  = seisansuu
        ws_daily.cell(row=row_b, column=4).value  = tensuu
        ws_daily.cell(row=row_b, column=6).value  = genuri1
        ws_daily.cell(row=row_b, column=8).value  = paypayuri
        ws_daily.cell(row=row_b, column=16).value = imanarauri
        ws_daily.cell(row=row_b, column=17).value = ichikauri
        ws_daily.cell(row=row_b, column=20).value = kinken
        ws_daily.cell(row=row_b, column=9).value  = credituri
        ws_daily.cell(row=row_b, column=13).value = koutuuri
        ws_daily.cell(row=row_b, column=19).value = uri13
        ws_daily.cell(row=row_b, column=18).value = uri11
        ws_daily.cell(row=row_b, column=15).value = nanacouri
        ws_daily.cell(row=row_b, column=14).value = waonuri
        ws_daily.cell(row=row_b, column=12).value = edyuri
        ws_daily.cell(row=row_b, column=11).value = iduri
        ws_daily.cell(row=row_b, column=10).value = ginrenuri

        if store_code != 7:  # 白浜(とれとれ市場)以外は別行にも転記
            ws_daily.cell(row=loop_idx + 73, column=3).value = genuri1

    # 全店合計を売上表へ転記 (col=2 が全体合計)
    uri01  = csv_val(csv2, 18,  2)  # 現金
    uri02  = csv_val(csv2, 25,  2)  # PayPay
    uri03  = csv_val(csv2, 115, 2)  # 商品券
    uri04  = csv_val(csv2, 28,  2)  # いまなら
    uri05  = csv_val(csv2, 31,  2)  # いちか
    uri06  = csv_val(csv2, 34,  2)  # クレジット
    uri07  = csv_val(csv2, 37,  2)  # ID
    uri40  = csv_val(csv2, 40,  2)  # 交通系
    uri43  = csv_val(csv2, 43,  2)  # Edy
    uri46  = csv_val(csv2, 46,  2)  # WAON
    uri49  = csv_val(csv2, 49,  2)  # nanaco
    uri52  = csv_val(csv2, 52,  2)  # 銀聯
    uri55  = csv_val(csv2, 55,  2)  # その他11
    uri58  = csv_val(csv2, 58,  2)  # その他12
    uri61  = csv_val(csv2, 61,  2)
    uri64  = csv_val(csv2, 64,  2)
    uri13t = csv_val(csv2, 79,  2)
    kyaku01 = csv_val(csv2, 6,  2)
    ten01   = csv_val(csv2, 8,  2)

    def _n(v): return v if isinstance(v, (int, float)) else 0
    uri_total = sum(_n(x) for x in [
        uri01, uri02, uri03, uri04, uri05, uri06, uri07,
        uri40, uri43, uri46, uri49, uri52, uri55, uri58, uri61, uri64, uri13t])

    # 売上表へ書き込み
    # tenkiCOL[0][1]+1=tenki_col, tenkiCOL[0][1]=tenki_col-1, tenkiCOL[0][1]-1=tenki_col-2
    xw(ws, excel_row, tenki_col,     weather)
    xw(ws, excel_row, tenki_col - 2, temp_max)
    xw(ws, excel_row, tenki_col - 1, temp_min)

    xw(ws, excel_row, uri_col,   uri_total)
    xw(ws, excel_row, kyaku_col, kyaku01)
    xw(ws, excel_row, tensu_col, ten01)

    # paypayCOL[0][1]+1=paypay_col, +2=paypay_col+1, ...
    xw(ws, excel_row, paypay_col,      uri02)  # PayPay
    xw(ws, excel_row, paypay_col + 1,  uri06)  # クレジット
    xw(ws, excel_row, paypay_col + 2,  uri07)  # ID
    xw(ws, excel_row, paypay_col + 3,  uri40)  # 交通系
    xw(ws, excel_row, paypay_col + 4,  uri43)  # Edy
    xw(ws, excel_row, paypay_col + 5,  uri46)  # WAON
    xw(ws, excel_row, paypay_col + 6,  uri49)  # nanaco
    xw(ws, excel_row, paypay_col + 7,  uri52)  # 銀聯
    xw(ws, excel_row, paypay_col + 8,  uri55)  # その他11

    if store_code == 6:  # 天理のみ追加列
        xw(ws, excel_row, paypay_col + 9,  uri05)  # いちか
        xw(ws, excel_row, paypay_col + 10, uri04)  # いまなら

    # 日報: 現金額を読む (F71) → 現金確認ダイアログへ
    cash_amount = ws_daily.cell(row=71, column=6).value
    wb_daily.save(str(daily_path))
    print(f"日報 保存: {daily_path.name}")

    # ── 追跡 Excel (直営売上) ─────────────────────────────────
    tracking_path = Path(cfg["tracking_dir"]) / f"{store_name}売上.xlsx"
    if tracking_path.exists():
        wb_track = openpyxl.load_workbook(str(tracking_path))
        ws_track = wb_track.worksheets[1]  # シートIndex=2 → 0-indexed=1
        free_row = 1
        for r in range(1, ws_track.max_row + 2):
            if ws_track.cell(row=r, column=1).value is None:
                free_row = r; break
        ws_track.cell(row=free_row, column=1).value = sale_date
        csv4 = read_csv_sj(filemei4)
        for i in range(1, len(csv4)):  # LoopIndex FROM 1 TO RowsCount-1
            val = csv4[i][1] if len(csv4[i]) > 1 else ""  # CSVTable4[i][1]
            try: val = int(str(val).replace(",",""))
            except Exception: pass
            ws_track.cell(row=free_row, column=i + 1).value = val
        wb_track.save(str(tracking_path))
        print(f"追跡Excel 保存: {tracking_path.name}")

    # ── 現金確認 ─────────────────────────────────────────────
    confirmed = dlg.show_cash_confirm(cash_amount)

    # ── 売上表を保存 ──────────────────────────────────────────
    wb_sales.save(str(sales_path))
    print(f"売上表 保存: {sales_path.name}")

    # ── メール送信 ────────────────────────────────────────────
    if confirmed:
        date_str   = f"{yr}{mm2}{dd2}"
        attach_path = Path(cfg["sales_table_attach"].replace("{store}", store_name))
        _send_email(cfg, date_str, str(attach_path if attach_path.exists() else sales_path), mm2, dd2)

    # ── CSV 削除 ──────────────────────────────────────────────
    for f in [filemei1, filemei2, filemei3, filemei4]:
        try: f.unlink(); print(f"削除: {f.name}")
        except FileNotFoundError: pass


def _send_email(cfg, date_str, attach_path, mm2, dd2):
    """Outlook でメール送信"""
    import subprocess
    try:
        subprocess.run(["taskkill", "/F", "/IM", "OUTLOOK.EXE"],
                       capture_output=True, timeout=10)
        time.sleep(10)
    except Exception:
        pass

    try:
        outlook = win32com.client.Dispatch("Outlook.Application")
        time.sleep(5)
        mail = outlook.CreateItem(0)

        for account in outlook.Session.Accounts:
            if cfg["email_from"] in (account.SmtpAddress or ""):
                mail.SendUsingAccount = account
                break

        mail.To      = cfg["email_to"]
        mail.CC      = cfg["email_cc"]
        mail.Subject = f"天理売上{date_str}"
        mail.Body    = (
            f"お疲れ様でございます。\n"
            f"{int(mm2)}月{int(dd2)}日の日報をお送りいたします。\n"
            f"ご確認ください。\n\n"
            f"☆☆☆☆☆☆☆☆☆☆☆☆☆☆☆☆☆☆☆☆☆\n"
            f"〒632-0073 奈良県 天理市 田町　410-1\n"
            f"お菓子のデパートよしや　天理店\n"
            f"℡/fax 0743-85-7236\n"
            f"tenri@okashinodepart.net\n\n"
            f"☆☆☆☆☆☆☆☆☆☆☆☆☆☆☆☆☆☆☆☆☆\n"
        )
        mail.Attachments.Add(attach_path)
        mail.Send()
        print("メール送信完了")
    except Exception as e:
        print(f"メール送信エラー: {e}")
        messagebox.showerror("メールエラー", f"メール送信に失敗しました:\n{e}")


# ─── エントリポイント ──────────────────────────────────────────
def main():
    cfg = load_config()
    dlg = Dialogs()

    try:
        store_name = dlg.ask_store(cfg["stores"])
        store_code = cfg["store_codes"][store_name]
        sale_date  = dlg.ask_date(date.today() - timedelta(days=1))
        operator   = dlg.ask_text("精算実行者名", "名前を入力してください")
        reg_counts = dlg.ask_registers(store_code)
        weather    = dlg.ask_text("天気入力", "天気を入力してください")
        temp_max   = dlg.ask_number("気温を入力", "最高気温を入力してください")
        temp_min   = dlg.ask_number("気温を入力", "最低気温を入力してください")

        print("=== Netdoa NX 接続・ダウンロード中... ===")
        download_csvs(cfg, store_name, store_code, sale_date)

        print("=== Excel 処理中... ===")
        process_seisan(
            dlg=dlg, cfg=cfg,
            store_name=store_name, store_code=store_code, sale_date=sale_date,
            operator=operator, reg_counts=reg_counts,
            weather=weather, temp_max=temp_max, temp_min=temp_min,
        )
        print("=== 完了 ===")

    except CancelError:
        print("キャンセルされました")
    except Exception as e:
        messagebox.showerror("エラー", f"処理中にエラーが発生しました:\n{e}")
        raise
    finally:
        dlg.destroy()


if __name__ == "__main__":
    main()
