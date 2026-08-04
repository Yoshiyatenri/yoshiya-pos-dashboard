"""
POSデータ 抽出ダッシュボード（SQLite / Supabase 自動切替）
- config.json の db_url が未設定 or XXXXXXXXXX → SQLiteで動作（ローカル）
- config.json の db_url が設定済み → Supabase（クラウド）
起動: streamlit run dashboard.py
"""
import io
import json
import re
import sqlite3
import calendar
from pathlib import Path
from datetime import datetime, timedelta, date

import pandas as pd
import streamlit as st

try:
    import psycopg2
    _psycopg2_ok = True
except ImportError:
    _psycopg2_ok = False

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    _openpyxl_ok = True
except ImportError:
    _openpyxl_ok = False

BASE_DIR = Path(__file__).parent
_cfg_path = BASE_DIR / "config.json"
cfg = json.loads(_cfg_path.read_text(encoding="utf-8")) if _cfg_path.exists() else {}

st.set_page_config(page_title="よしや POSデータ抽出", layout="wide")


def check_password() -> bool:
    """パスワード認証（通常ユーザー / 管理者を区別）"""
    try:
        correct = st.secrets["password"]
        admin_pw = st.secrets["admin_password"]
    except Exception:
        correct = cfg.get("dashboard_password", "")
        admin_pw = cfg.get("admin_password", "")

    if st.session_state.get("authenticated"):
        return True

    st.title("📊 よしや POSデータ抽出ダッシュボード")
    pw = st.text_input("パスワードを入力してください", type="password", key="pw_input")
    if pw:
        if pw == admin_pw:
            st.session_state["authenticated"] = True
            st.session_state["is_admin"] = True
            st.rerun()
        elif pw == correct:
            st.session_state["authenticated"] = True
            st.session_state["is_admin"] = False
            st.rerun()
        else:
            st.error("パスワードが違います")
    return False


if not check_password():
    st.stop()


def save_access_log(stores: list, start: str, end: str, jan: str, cats: list) -> None:
    """抽出操作をaccess_logsテーブルに保存（Supabase接続時のみ）"""
    if not _use_pg():
        return
    try:
        headers = st.context.headers
        ip = headers.get("X-Forwarded-For", headers.get("X-Real-Ip", "不明"))
        # X-Forwarded-Forは複数IPが含まれる場合があるので先頭のみ取得
        ip = ip.split(",")[0].strip() if ip != "不明" else "不明"
        user_agent = headers.get("User-Agent", "不明")

        con = get_conn()
        cur = con.cursor()
        cur.execute(
            "INSERT INTO access_logs (stores, start_date, end_date, jan_code, categories, ip_address, user_agent) VALUES (%s, %s, %s, %s, %s, %s, %s)",
            (", ".join(stores), start, end, jan.strip(), ", ".join(cats), ip, user_agent),
        )
        con.commit()
        cur.close()
        con.close()
    except Exception:
        pass

st.title("📊 よしや POSデータ抽出ダッシュボード")

PREFIX_RE = re.compile(r'^お菓子のデパート\s*よしや\s*')


def shorten(name: str) -> str:
    return PREFIX_RE.sub("", name)


def get_db_url() -> str:
    """接続URL: Streamlit Secrets優先、なければconfig.json"""
    try:
        return st.secrets["database"]["url"]
    except Exception:
        return cfg.get("db_url", "")


def _use_pg() -> bool:
    """Supabaseを使うかどうか（URLが正しく設定されている場合のみ）"""
    url = get_db_url()
    return _psycopg2_ok and bool(url) and "XXXXXXXXXX" not in url


def get_conn():
    """DB接続を返す（SQLite / Supabase 自動切替）"""
    if _use_pg():
        return psycopg2.connect(get_db_url())
    sqlite_path = BASE_DIR / cfg.get("db_path_sqlite", "pos_data.db")
    return sqlite3.connect(str(sqlite_path))


def _ph(n: int) -> str:
    """プレースホルダーをn個生成（SQLite:? / PostgreSQL:%s）"""
    mark = "%s" if _use_pg() else "?"
    return ",".join([mark] * n)


def _fix(sql: str) -> str:
    """SQLite用にプレースホルダーを変換（%s → ?）"""
    if _use_pg():
        return sql
    return sql.replace("%s", "?")


def get_distinct(column: str, where_extra: str = "", params_extra: list | None = None) -> list:
    """指定カラムの重複なし値一覧（フィルター条件追加可）"""
    params_extra = params_extra or []
    where = f"WHERE {column} IS NOT NULL AND {column} != ''"
    if where_extra:
        where += f" AND {where_extra}"
    try:
        con = get_conn()
        cur = con.cursor()
        cur.execute(
            _fix(f"SELECT DISTINCT {column} FROM sales {where} ORDER BY {column}"),
            params_extra,
        )
        rows = cur.fetchall()
        cur.close()
        con.close()
        return [r[0] for r in rows]
    except Exception:
        return []


@st.cache_data(ttl=300)
def get_store_list() -> tuple[list[str], dict[str, str]]:
    """(表示名リスト, 表示名→DB名マッピング)"""
    try:
        con = get_conn()
        cur = con.cursor()
        cur.execute("SELECT DISTINCT store_name FROM sales ORDER BY store_name")
        rows = cur.fetchall()
        cur.close()
        con.close()
    except Exception:
        return [], {}
    full_names = [r[0] for r in rows]
    display = [shorten(n) for n in full_names]
    return display, dict(zip(display, full_names))


@st.cache_data(ttl=300)
def get_months() -> list[str]:
    """DBに存在する月一覧（降順）"""
    try:
        con = get_conn()
        cur = con.cursor()
        cur.execute("SELECT DISTINCT substr(pos_date, 1, 7) FROM sales ORDER BY 1 DESC")
        rows = cur.fetchall()
        cur.close()
        con.close()
        return [r[0] for r in rows]
    except Exception:
        return []


def query_data(
    start: str,
    end: str,
    stores_db: list[str],
    jan: str,
    cats: list[str],
    mid_cats: list[str],
    small_cats: list[str],
) -> pd.DataFrame:
    """フィルター条件でデータ取得"""
    where = f"pos_date BETWEEN {_ph(1)} AND {_ph(1)}"
    params: list = [start, end]
    if stores_db:
        where += f" AND store_name IN ({_ph(len(stores_db))})"
        params.extend(stores_db)
    if jan.strip():
        where += f" AND plu_code = {_ph(1)}"
        params.append(jan.strip())
    if cats:
        where += f" AND cat_name IN ({_ph(len(cats))})"
        params.extend(cats)
    if mid_cats:
        where += f" AND mid_cat_name IN ({_ph(len(mid_cats))})"
        params.extend(mid_cats)
    if small_cats:
        where += f" AND small_cat_name IN ({_ph(len(small_cats))})"
        params.extend(small_cats)
    try:
        con = get_conn()
        df = pd.read_sql_query(
            f"""
            SELECT pos_date, store_name,
                   COALESCE(cat_name, '') AS cat_name,
                   COALESCE(mid_cat_name, '') AS mid_cat_name,
                   COALESCE(small_cat_name, '') AS small_cat_name,
                   plu_code, plu_name, master_cost, master_price,
                   sales_amount, sales_qty, sales_customers, gross_profit
            FROM sales
            WHERE {where}
            ORDER BY pos_date, store_name, plu_code
            """,
            con,
            params=params,
        )
        con.close()
        return df
    except Exception as e:
        st.error(f"データ取得エラー: {e}")
        return pd.DataFrame()


def query_bumon_analysis(start: str, end: str, stores_db: list[str]) -> pd.DataFrame:
    """small_cat_name 別の部門別分析を返す"""
    where = f"pos_date BETWEEN {_ph(1)} AND {_ph(1)} AND small_cat_name IS NOT NULL AND small_cat_name != ''"
    params: list = [start, end]
    if stores_db:
        where += f" AND store_name IN ({_ph(len(stores_db))})"
        params.extend(stores_db)
    sql = f"""
        SELECT
            small_cat_name AS cat,
            SUM(sales_qty)    AS qty,
            SUM(sales_amount) AS amount,
            SUM(gross_profit) AS profit,
            COUNT(DISTINCT plu_code) AS sku_count
        FROM sales
        WHERE {where}
        GROUP BY small_cat_name
        ORDER BY SUM(sales_amount) DESC
    """
    try:
        con = get_conn()
        df = pd.read_sql_query(_fix(sql), con, params=params)
        con.close()
    except Exception as e:
        st.error(f"部門別分析エラー: {e}")
        return pd.DataFrame()
    if df.empty:
        return df

    df = df.rename(columns={"cat": "カテゴリー", "qty": "売数", "amount": "売上",
                             "profit": "荒利", "sku_count": "SKU数"})
    df["荒利率"]    = (df["荒利"] / df["売上"]).where(df["売上"] > 0)
    df["単価"]      = (df["売上"] / df["売数"]).where(df["売数"] > 0)
    df["1SKU当売数"] = (df["売数"] / df["SKU数"]).where(df["SKU数"] > 0)
    df["1SKU当売上"] = (df["売上"] / df["SKU数"]).where(df["SKU数"] > 0)
    df["1SKU当荒利"] = (df["荒利"] / df["SKU数"]).where(df["SKU数"] > 0)
    df["売上順位"]   = range(1, len(df) + 1)
    df["荒利率順位"] = df["荒利率"].rank(ascending=False, method="min").astype("Int64")

    total_売上  = df["売上"].sum()
    total_荒利  = df["荒利"].sum()
    total_売数  = df["売数"].sum()
    total = pd.DataFrame([{
        "カテゴリー": "合計",
        "売数": total_売数,
        "売上": total_売上,
        "荒利": total_荒利,
        "SKU数": df["SKU数"].sum(),
        "荒利率": total_荒利 / total_売上 if total_売上 > 0 else None,
        "単価":   total_売上 / total_売数 if total_売数 > 0 else None,
        "1SKU当売数": None, "1SKU当売上": None, "1SKU当荒利": None,
        "売上順位": None, "荒利率順位": None,
    }])
    return pd.concat([df, total], ignore_index=True)


def make_bumon_excel(df: pd.DataFrame, store_label: str, start: str, end: str) -> bytes | None:
    """部門別分析DataFrameをExcelに変換してbytesで返す"""
    if not _openpyxl_ok:
        return None
    wb = Workbook()
    ws = wb.active
    ws.title = "部門別集計"

    # タイトル行（期間を反映）
    if start[:7] == end[:7]:
        y, m = start[:4], str(int(start[5:7]))
        period_label = f"{y}年{m}月"
    else:
        period_label = f"{start}〜{end}"
    ws["A1"] = f"{store_label}　{period_label}　部門別実績（POSデータ集計）"
    ws["A1"].font = Font(bold=True, size=12)

    # ヘッダー行（3行目）
    headers = ["カテゴリー", "売数", "売上", "荒利", "荒利率", "SKU数",
               "単価", "1SKU当売数", "1SKU当売上", "1SKU当荒利", "売上順位", "荒利率順位"]
    hfill = PatternFill("solid", fgColor="4472C4")
    hfont = Font(bold=True, color="FFFFFF")
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=3, column=ci, value=h)
        c.fill = hfill
        c.font = hfont
        c.alignment = Alignment(horizontal="center")

    # データ行（4行目〜）
    num_fmts = ["@", "#,##0", "#,##0", "#,##0.00", "0.0%",
                "#,##0", "#,##0.000", "#,##0.000", "#,##0.00", "#,##0.000", "0", "0"]
    for ri, (_, row) in enumerate(df.iterrows(), 4):
        is_total = row["カテゴリー"] == "合計"
        def _v(key, cast=float):
            val = row.get(key)
            return cast(val) if pd.notna(val) else None
        vals = [
            row["カテゴリー"],
            _v("売数", int),
            _v("売上", int),
            _v("荒利"),
            _v("荒利率"),
            _v("SKU数", int),
            _v("単価"),
            _v("1SKU当売数"),
            _v("1SKU当売上"),
            _v("1SKU当荒利"),
            _v("売上順位", int),
            _v("荒利率順位", int),
        ]
        for ci, (v, fmt) in enumerate(zip(vals, num_fmts), 1):
            cell = ws.cell(row=ri, column=ci, value=v)
            cell.number_format = fmt
            if is_total:
                cell.font = Font(bold=True)

    # 列幅
    for ci, w in enumerate([14, 10, 12, 12, 8, 8, 12, 12, 14, 14, 8, 10], 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─── 接続チェック ─────────────────────────────────────────────────────────────
if _use_pg():
    st.sidebar.success("☁️ Supabase（クラウド）接続中")
else:
    sqlite_path = BASE_DIR / cfg.get("db_path_sqlite", "pos_data.db")
    if not sqlite_path.exists():
        st.warning("データベースがまだありません。先に `run_daily.bat` を実行してください。")
        st.stop()
    st.sidebar.info("📂 ローカルSQLiteで動作中")

# ─── サイドバー（抽出条件） ──────────────────────────────────────────────────
with st.sidebar:
    st.header("🔍 抽出条件")

    # ① 日付
    st.subheader("① 日付")
    date_mode = st.radio("入力方法", ["日付範囲", "月選択", "月範囲"], horizontal=True)
    today = datetime.now().date()

    if date_mode == "日付範囲":
        start_date = st.date_input("開始日", today - timedelta(days=1))
        end_date = st.date_input("終了日", today - timedelta(days=1))

    elif date_mode == "月選択":
        months = get_months()
        if not months:
            st.warning("データがありません。")
            st.stop()
        selected_month = st.selectbox("月", months)
        y, m = map(int, selected_month.split("-"))
        start_date = date(y, m, 1)
        end_date = date(y, m, calendar.monthrange(y, m)[1])
        st.caption(f"期間: {start_date} 〜 {end_date}")

    else:  # 月範囲
        months = get_months()
        if not months:
            st.warning("データがありません。")
            st.stop()
        col_s, col_e = st.columns(2)
        with col_s:
            start_month = st.selectbox("開始月", months, index=len(months) - 1)
        with col_e:
            end_month = st.selectbox("終了月", months, index=0)
        y_s, m_s = map(int, start_month.split("-"))
        y_e, m_e = map(int, end_month.split("-"))
        start_date = date(y_s, m_s, 1)
        end_date = date(y_e, m_e, calendar.monthrange(y_e, m_e)[1])
        if start_date > end_date:
            st.error("開始月が終了月より後になっています。")
            st.stop()
        st.caption(f"期間: {start_date} 〜 {end_date}")

    # ② 店舗
    st.subheader("② 店舗")
    display_stores, store_mapping = get_store_list()
    selected_display = st.multiselect("店舗名（複数選択可）", display_stores, default=display_stores)
    selected_stores_db = [store_mapping[d] for d in selected_display]

    # ③ JANコード / PLUコード
    st.subheader("③ JANコード / PLUコード")
    jan_input = st.text_input("（空白で全件）", "")

    # ④ 分類（連動フィルター）
    st.subheader("④ 分類（未選択で全件）")

    all_cats = get_distinct("cat_name")
    selected_cats = st.multiselect("カテゴリー", all_cats)

    # 中分類：カテゴリー選択に連動
    if selected_cats:
        all_mid_cats = get_distinct(
            "mid_cat_name",
            f"cat_name IN ({_ph(len(selected_cats))})",
            selected_cats,
        )
    else:
        all_mid_cats = get_distinct("mid_cat_name")
    selected_mid_cats = st.multiselect("中分類", all_mid_cats)

    # 小分類：カテゴリー＋中分類選択に連動
    ex_where_parts, ex_params = [], []
    if selected_cats:
        ex_where_parts.append(f"cat_name IN ({_ph(len(selected_cats))})")
        ex_params.extend(selected_cats)
    if selected_mid_cats:
        ex_where_parts.append(f"mid_cat_name IN ({_ph(len(selected_mid_cats))})")
        ex_params.extend(selected_mid_cats)
    all_small_cats = get_distinct(
        "small_cat_name",
        " AND ".join(ex_where_parts),
        ex_params or None,
    )
    selected_small_cats = st.multiselect("小分類", all_small_cats)

    # ⑤ 集計オプション
    st.subheader("⑤ 集計オプション")
    aggregate_period = st.checkbox("期間合計（日別にせず集計する）", value=False)
    store_agg = st.radio("店舗の集計方法", ["店別に集計", "選択店舗の合計"])

    st.divider()
    extract_btn = st.button("🔍 抽出実行", use_container_width=True, type="primary")
    bumon_btn   = st.button("📊 部門別分析レポート", use_container_width=True)

# ─── 抽出処理 ────────────────────────────────────────────────────────────────
if extract_btn:
    if not selected_display:
        st.warning("店舗を1つ以上選択してください。")
        st.stop()

    with st.spinner("データを取得中..."):
        df_raw = query_data(
            str(start_date), str(end_date),
            selected_stores_db, jan_input,
            selected_cats, selected_mid_cats, selected_small_cats,
        )
        save_access_log(selected_display, str(start_date), str(end_date), jan_input, selected_cats)

    if df_raw.empty:
        st.warning("該当データがありません。")
        st.session_state.pop("df_result", None)
    else:
        df_raw["store_name"] = df_raw["store_name"].apply(shorten)

        # グループキーの決定
        # JAN指定+店別集計の場合は日付を集計しない（店ごとに期間合計で1行表示）
        jan_specified = bool(jan_input.strip())
        use_period_agg = aggregate_period or (jan_specified and store_agg == "店別に集計")

        group_keys = ["cat_name", "mid_cat_name", "small_cat_name", "plu_code", "plu_name"]
        if not use_period_agg:
            group_keys = ["pos_date"] + group_keys
        if store_agg == "店別に集計":
            group_keys = ["store_name"] + group_keys

        df_agg = (
            df_raw
            .groupby(group_keys, as_index=False)
            .agg(
                master_cost=("master_cost", "first"),
                master_price=("master_price", "first"),
                sales_amount=("sales_amount", "sum"),
                sales_qty=("sales_qty", "sum"),
                sales_customers=("sales_customers", "sum"),
                gross_profit=("gross_profit", "sum"),
            )
        )

        # POS日付1（開始）・日付2（終了）を先頭に付与
        df_agg.insert(0, "POS日付2（終了）", str(end_date))
        df_agg.insert(0, "POS日付1（開始）", str(start_date))

        df_agg = df_agg.rename(columns={
            "pos_date": "日付",
            "store_name": "店舗名",
            "cat_name": "カテゴリー",
            "mid_cat_name": "中分類",
            "small_cat_name": "小分類",
            "plu_code": "PLUコード",
            "plu_name": "PLU名",
            "master_cost": "マスター原価",
            "master_price": "マスター売価",
            "sales_amount": "売上",
            "sales_qty": "点数",
            "sales_customers": "客数",
            "gross_profit": "粗利",
        })

        st.session_state["df_result"] = df_agg
        st.session_state["totals"] = {
            "sales": df_raw["sales_amount"].sum(),
            "profit": df_raw["gross_profit"].sum(),
        }
        st.session_state["period"] = (str(start_date), str(end_date))

if bumon_btn:
    if not selected_display:
        st.warning("店舗を1つ以上選択してください。")
    else:
        with st.spinner("部門別データを集計中..."):
            df_bumon = query_bumon_analysis(str(start_date), str(end_date), selected_stores_db)
        if df_bumon.empty:
            st.warning("該当データがありません。")
        else:
            st.session_state["bumon_result"] = df_bumon
            st.session_state["bumon_meta"] = {
                "stores": selected_display,
                "start": str(start_date),
                "end": str(end_date),
            }

if not extract_btn and "df_result" not in st.session_state and "bumon_result" not in st.session_state:
    st.info("👈 左のサイドバーで条件を設定して「抽出実行」または「部門別分析レポート」ボタンを押してください。")
    st.stop()

# ─── 結果表示 ────────────────────────────────────────────────────────────────
if "df_result" in st.session_state:
    df_show = st.session_state["df_result"]
    totals = st.session_state["totals"]
    period = st.session_state["period"]

    st.subheader(f"📋 抽出結果　{period[0]} 〜 {period[1]}")

    c1, c2, c3, c4 = st.columns(4)
    c1.metric("抽出行数", f"{len(df_show):,} 行")
    c2.metric("売上合計", f"¥{totals['sales']:,.0f}")
    c3.metric("粗利合計", f"¥{totals['profit']:,.0f}")
    gross_rate = totals["profit"] / totals["sales"] * 100 if totals["sales"] > 0 else 0
    c4.metric("粗利率", f"{gross_rate:.1f}%")

    st.divider()
    st.dataframe(df_show, use_container_width=True, hide_index=True)

    csv_bytes = df_show.to_csv(index=False, encoding="utf-8-sig").encode("utf-8-sig")
    st.download_button(
        label="📥 CSV ダウンロード",
        data=csv_bytes,
        file_name=f"pos_extract_{period[0]}_{period[1]}.csv",
        mime="text/csv",
    )

# ─── 部門別分析レポート ───────────────────────────────────────────────────────
if "bumon_result" in st.session_state:
    df_b = st.session_state["bumon_result"]
    meta = st.session_state["bumon_meta"]
    store_label = "・".join(meta["stores"])
    st.divider()
    st.subheader(f"📊 部門別分析　{meta['start']} 〜 {meta['end']}")
    st.caption(f"店舗: {store_label}")

    # 表示用（荒利率を%表示）
    df_disp = df_b.copy()
    df_disp["荒利率"] = df_disp["荒利率"].apply(
        lambda v: f"{v:.1%}" if pd.notna(v) else ""
    )
    st.dataframe(df_disp, use_container_width=True, hide_index=True)

    if _openpyxl_ok:
        excel_bytes = make_bumon_excel(df_b, store_label, meta["start"], meta["end"])
        if excel_bytes:
            if meta["start"][:7] == meta["end"][:7]:
                period_str = meta["start"][:7]
            else:
                period_str = f"{meta['start']}_{meta['end']}"
            fname = f"部門別分析_{store_label}_{period_str}.xlsx"
            st.download_button(
                label="📥 Excelダウンロード",
                data=excel_bytes,
                file_name=fname,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
    else:
        st.warning("openpyxl がインストールされていません。`pip install openpyxl` を実行してください。")

# ─── 管理者：アクセスログ ─────────────────────────────────────────────────────
if st.session_state.get("is_admin") and _use_pg():
    st.divider()
    st.subheader("🔐 アクセスログ（管理者専用）")
    try:
        con = get_conn()
        df_logs = pd.read_sql_query(
            "SELECT accessed_at, stores, start_date, end_date, jan_code, categories, ip_address, user_agent FROM access_logs ORDER BY accessed_at DESC LIMIT 200",
            con,
        )
        con.close()
        df_logs["accessed_at"] = (
            pd.to_datetime(df_logs["accessed_at"], utc=True)
            .dt.tz_convert("Asia/Tokyo")
            .dt.strftime("%Y-%m-%d %H:%M:%S")
        )
        df_logs = df_logs.rename(columns={
            "accessed_at": "日時（JST）",
            "stores": "店舗",
            "start_date": "開始日",
            "end_date": "終了日",
            "jan_code": "JAN",
            "categories": "カテゴリー",
            "ip_address": "IPアドレス",
            "user_agent": "ブラウザ",
        })
        st.dataframe(df_logs, use_container_width=True, hide_index=True)
    except Exception as e:
        st.error(f"ログ取得エラー: {e}")
