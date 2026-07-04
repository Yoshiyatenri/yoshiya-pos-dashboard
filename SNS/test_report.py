from openpyxl import load_workbook
from report import create_report


def test_create_report_writes_trend_and_plan_sheets(tmp_path):
    summary = {
        "video_count": 2,
        "top_hashtags": [("駄菓子", 2), ("うまい棒", 1)],
        "top_videos_by_likes": [
            {
                "caption": "駄菓子屋開封動画",
                "likes": 300,
                "posted_at": "2026-06-05",
                "url": "https://www.tiktok.com/@b/video/2",
            },
        ],
    }
    plan_ideas = ["うまい棒の食べ比べ動画を作る"]
    output_path = tmp_path / "report.xlsx"

    create_report(summary, plan_ideas, str(output_path))

    wb = load_workbook(output_path)
    assert wb.sheetnames == ["傾向データ", "投稿企画案"]

    trend_sheet = wb["傾向データ"]
    assert trend_sheet["A1"].value == "集計投稿数"
    assert trend_sheet["B1"].value == 2
    assert trend_sheet["C8"].value == "2026-06-05"

    plan_sheet = wb["投稿企画案"]
    assert plan_sheet["A2"].value == "うまい棒の食べ比べ動画を作る"
